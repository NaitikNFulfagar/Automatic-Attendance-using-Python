import cv2 #READ FROM CAMERAa
import numpy as np
import pyttsx3 #AUDIO PLAY OF PRESENTY
from pyzbar.pyzbar import decode #QRCODEDECODE
from datetime import datetime
import time
import openpyxl   #EXCEL EXPORT
from openpyxl import load_workbook
import os

x = datetime.now()


cap = cv2.VideoCapture(0)#Capture Via Camera.
cap.set(3,640) # Width Of Camera
cap.set(4,480) # Height Of Camera
rollnop=[]
roll={
1: "Ayushi Bhandari",
2: "Gayatri Chaudhari",
3: "Srushti Chaudhary",
4: "Sujal Deore",
5: "Pawan Desale",
6: "Hemant Deshmukh",
7: "Vaishnavi Dusane",
8: "Naitik Fulfagar",
9: "Saee Gadhekar",
10: "Payal Gaikwad",
11: "Vaibhavi Gangurde",
12: "Aditya Jadhav",
13: "Rahul Jadhav",
14: "Ujjwal Katare",
15: "Cancelled",
16: "Priti Khaire",
17: "Tanmay Khairnar",
18: "Rutuja Lomate",
19: "Ashwini Mali",
20: "Amey Mohole",
21: "Parth Pagar",
22: "Durgesh Pagar",
23: "Keshavdas Panpatil",
24: "Mayuri Pardeshi",
25: "Ritesh Pardeshi",
26: "Tanvi Pardeshi",
27: "Tejas Patil",
28: "Vishnavi Patil",
29: "Tilak Rokaya",
30: "Sudarshan Sanap",
31: "Taha Shaikh",
32: "Ajinkya Shinde",
33: "Sayee Shirsath",
34: "Aditya Sonawane",
35: "Sumedh Takalkar",
36: "Varad Tannu",
37: "Pranav Thorat",
38: "Aditya Tokare",
39: "Rutuja Vaidya",
40: "Aashish Wagh",
41: "Prajwal Wankhede",
42: "Prerana Patil",
43: "Ruchika Kapadnis",
      }
#------------------------------------------------------
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[0].id) #changing index changes voices but ony 0 and 1 are working here
engine.setProperty('rate', 150)     # setting up new voice rate
engine.setProperty('volume',1.0)    # setting up volume level  between 0 and 1
#--------------------------------------------------------------

#def exc(loc):

#--------------------------------------------
dirName=x.strftime("%d") + "-" + x.strftime("%m") + "-" + x.strftime("%Y")
if not os.path.exists(dirName):
    os.mkdir(dirName)
  #  print("Directory " , dirName ,  " Created ")
#------------------------------------------------

exce=1
wb = openpyxl.Workbook()
sheet = wb.active

while(exce<44):
        
        sheet['A1'] = "Roll No.:"
        sheet['B1'] = "Name:"
        sheet['C1'] = "In Time:"
        sheet['D1'] = x.strftime("%x")
        sheet['E1'] = "Present Students - "
        sheet['H1'] = "Sequence Of Student Attendance"
        #sheet['A4'] = "Roll No.:"
        c1 = sheet.cell(row= exce+1 , column = 1) 
        c1.value = exce
        c2 = sheet.cell(row= exce+1 , column = 2)
        c2.value = (roll[exce])
        #c3 = sheet.cell(row= sd+1   , column = 3)
        #c3.value = now
        exce=exce+1
        wb.save(dirName+"/Attendance.xlsx")
        

#---------------------------------------------------------------
now=str(datetime.now())


def decoder(image):
    gray_img = cv2.cvtColor(image,0)    #Change Image In backend For Quicker Responce or Quicker Getting scan
    barcode = decode(gray_img)

    for obj in barcode:
           
        points = obj.polygon
        (x,y,w,h) = obj.rect
        pts = np.array(points, np.int32)
        pts = pts.reshape((-1, 1, 2))
        cv2.polylines(image, [pts], True, (0, 255, 0), 3)

        barcodeData = obj.data.decode("utf-8")
        barcodeType = obj.type
        barcodeDataint = int(barcodeData)
        string = "Hi Roll no. " + str(barcodeData)
        st1= now
        #exc(barcodeData)
       
        cv2.putText(frame, string, (x,y), cv2.FONT_HERSHEY_SIMPLEX,0.8,(255,0,0), 2)
        file1 = open(dirName+"/attendance.txt","a")
        L = ["Roll No."+barcodeData+" "]
        if barcodeDataint  in roll and not barcodeDataint in rollnop :
                    name=str(roll[barcodeDataint])
                    print("Hello Roll No."+barcodeData+" "+name+now)
                    text = ("Roll Number"+barcodeData+" "+name+" Present")
                    engine.say(text)
                    engine.runAndWait()
                    #time.sleep(5)
                
        
                # \n is placed to indicate EOL (End of Line)
        #if barcodeDataint in roll and not in:
                    file1.write("\n")
                    file1.writelines(L)
                    file1.write(now)
                    file1.close()
                    rollnop.append(barcodeDataint)
                   # sheet['E1'] = len(rollnop)
                    #--------------------------------
                    workbook = load_workbook(filename=dirName+"/Attendance.xlsx")
                    sheet = workbook.active
                    sd=int(barcodeData)
                    cm = sheet.cell(row= sd+1 , column = 3)
                    nai = datetime.now()
                    cm.value = nai.strftime("%X")
                    sheet['E2'] = len(rollnop)
                    listToStr = ' ,'.join([str(elem) for elem in rollnop])
                    cm1 = sheet.cell(row= jio+2 , column = 8)
                    cm1.value = listToStr
                    workbook.save(filename=dirName+"/Attendance.xlsx")
                     #--------------------------------   
        elif barcodeDataint in rollnop:
                    tex1 = " Duplicate Entry"
                    print(tex1)
                    engine.say(tex1)
                    engine.runAndWait()    
        else:
                    text = " Invaild Input "
                    print(text)
                    engine.say(text)
                    engine.runAndWait()
                   # time.sleep(5)
            
        

while True:
    jio=0
    ret, frame = cap.read()
    decoder(frame)
    jio=jio+1
    cv2.imshow('Attendance Scanner', frame)
    code = cv2.waitKey(1)
    
    if code == ord('q'):
        break



